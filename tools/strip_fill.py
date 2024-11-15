#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import os
import openpyxl


# In[19]:


def read_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames
    second_sheet_name = sheet_names[2]
    sheet = workbook[second_sheet_name]

    # 逐行读取数据并保存到列表
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # 将数据转换为DataFrame
    df = pd.DataFrame(data)
    return df


# In[4]:


def get_invest_type(df):
    invest_type = []
    for index, value in df[1].items():
        if isinstance(value, str) and value.endswith("类"):
            invest_type.append((index, value))
    return invest_type


# In[16]:


def process_rows(df, invest_type):
    new_rows = []
    
    for i in range(len(invest_type)):
        start_index, invest_type_value = invest_type[i]
        end_index = invest_type[i + 1][0] if i + 1 < len(invest_type) else len(df)
        
        for j in range(start_index, end_index):
            row = df.iloc[j]
            if pd.notna(row[1]) and pd.notna(row[2]):
                # Extract numeric part from column 3
                scale = ''.join(filter(str.isdigit, str(row[3])))
                # Convert column 4 to datetime
                start_time = pd.to_datetime(row[4], format='%Y', errors='coerce')
                new_row = [row[1], invest_type_value, row[2], scale, start_time]
                new_rows.append(new_row)
    new_df = pd.DataFrame(new_rows, columns=['proj_name', 'proj_phase', 'proj_content', 'scale', 'start_time'])
    return new_df


# In[ ]:


def process_all_files(directory):
    result = pd.DataFrame()
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            # 生成pdf文件的路径
            file_path = os.path.join(directory, filename)# 读取xlsx文件并存入DataFrame
            df = read_xlsx(file_path)
            invest_type = get_invest_type(df)
            result = pd.concat([result, process_rows(df, invest_type)], ignore_index=True)
    return result

